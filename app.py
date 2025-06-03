import streamlit as st
import tempfile
import os
from pdf_handler import PDFHandler
from ocr_processor import OCRProcessor
from translator import Translator
import traceback
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_excel_file(df_table_data):
    """
    Create Excel file from a Pandas DataFrame with proper formatting.
    """
    if df_table_data is None or df_table_data.empty:
        st.warning("No data to create Excel file.")
        # Return an empty BytesIO object or handle as an error if preferred
        return BytesIO() 

    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_table_data.to_excel(writer, index=False, sheet_name='Polling Stations')
        workbook = writer.book
        worksheet = writer.sheets['Polling Stations']
        
        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass # Handle cases where len() might fail (e.g. on numbers)
            adjusted_width = (max_length + 2) * 1.2
            # Cap width to a reasonable maximum, e.g., 100
            adjusted_width = min(adjusted_width, 100)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col_num, value in enumerate(df_table_data.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            # Value is already set by df.to_excel, we just format it
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Apply borders and text wrapping to all data cells (excluding header)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df_table_data) + 1, max_col=len(df_table_data.columns)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1: # Apply to data rows, not header
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                else: # Header alignment is already set
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set a default row height for data rows for better readability if text wraps
        for i in range(2, len(df_table_data) + 2):
            worksheet.row_dimensions[i].height = 30 # Adjust as needed
            
    output.seek(0)
    return output

def process_single_pdf(pdf_file, pdf_handler, ocr_processor, translator):
    """
    Process a single PDF file
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
        tmp_pdf.write(pdf_file.getvalue())
        tmp_pdf_path = tmp_pdf.name
    
    try:
        images = pdf_handler.pdf_to_images(tmp_pdf_path)
        if not images:
            st.error("Could not convert PDF to images.")
            return None, None, None
        
        all_punjabi_text_parts = []
        for img_idx, img in enumerate(images):
            # st.write(f"Processing image {img_idx + 1}/{len(images)}...") # Optional: for debugging page-by-page progress
            text_part = ocr_processor.extract_punjabi_text(img)
            if text_part: # Ensure we don't add empty strings if OCR fails for a page
                all_punjabi_text_parts.append(text_part)
        
        punjabi_text = "\n\n".join(all_punjabi_text_parts) # Join text from all pages, using double newline for better separation
        
        if not punjabi_text:
            st.error("OCR processing failed to extract text from any page.")
            return None, None, None
            
        translated_text = translator.translate_to_english(punjabi_text)
        if not translated_text or "Translation error:" in translated_text:
            st.error(f"Translation failed: {translated_text}")
            return punjabi_text, translated_text, None # Return translated_text for display even if table fails
        
        # Format translated text into DataFrame
        df_table_data = translator.format_translated_text_to_dataframe(translated_text)
        
        return punjabi_text, translated_text, df_table_data
    finally:
        os.remove(tmp_pdf_path)

def main():
    st.set_page_config(
        page_title="Punjabi PDF OCR & Translation",
        page_icon="📄",
        layout="wide"
    )
    
    st.title("📄 Punjabi PDF OCR & Translation")
    st.markdown("Extract Punjabi text from PDFs and translate to English with Excel export")
    
    # Initialize processors
    pdf_handler = PDFHandler()
    ocr_processor = OCRProcessor()
    translator = Translator()
    
    # Processing mode selection
    processing_mode = st.radio(
        "Choose processing mode:",
        ["Single PDF", "Batch Processing"],
        horizontal=True
    )
    
    if processing_mode == "Single PDF":
        # Single file upload
        uploaded_file = st.file_uploader(
            "Choose a PDF file",
            type="pdf",
            help="Upload a PDF file containing Punjabi text"
        )
        
        if uploaded_file is not None:
            # Display file info
            st.success(f"✅ File uploaded: {uploaded_file.name} ({uploaded_file.size} bytes)")
            
            # Process button
            if st.button("🔄 Process PDF", type="primary"):
                try:
                    # Create progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    status_text.text("📖 Processing PDF...")
                    progress_bar.progress(20)
                    
                    punjabi_text, translated_text, df_table_data = process_single_pdf(
                        uploaded_file, pdf_handler, ocr_processor, translator
                    )
                    
                    if punjabi_text:
                        st.subheader("📄 Original Punjabi Text")
                        st.text_area("Original Text", value=punjabi_text, height=300, key="original_single")

                    if translated_text:
                        st.subheader("📜 Translated English Text")
                        st.text_area("Translated Text", value=translated_text, height=300, key="translated_single")
                        st.download_button(
                            label="📝 Download Translated Text",
                            data=translated_text,
                            file_name=f"{uploaded_file.name}_translated.txt",
                            mime="text/plain",
                            key="download_translated_single"
                        )

                    if df_table_data is not None and not df_table_data.empty:
                        st.subheader("📊 Extracted Polling Station Data")
                        st.dataframe(df_table_data)
                        
                        excel_data = create_excel_file(df_table_data) # create_excel_file now expects a DataFrame
                        st.download_button(
                            label="📊 Download as Excel",
                            data=excel_data,
                            file_name=f"{uploaded_file.name}_polling_stations.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_single"
                        )
                    elif translated_text:
                        st.warning("Could not extract structured table data from the translated text, or the table is empty. Please review the translated text.")
                    
                    if not punjabi_text and not translated_text:
                        st.error("Processing failed completely. Please check the PDF file and try again.")
                        
                except Exception as e:
                    st.error(f"❌ An error occurred during processing: {str(e)}")
                    st.error("Please ensure your PDF contains readable Punjabi text and try again.")
                    # Show detailed error in expander for debugging
                    with st.expander("Show detailed error"):
                        st.code(traceback.format_exc())
    
    else:  # Batch Processing
        st.markdown("### Batch Processing Mode")
        uploaded_files = st.file_uploader(
            "Choose multiple PDF files",
            type="pdf",
            accept_multiple_files=True,
            help="Upload multiple PDF files containing Punjabi text"
        )
        
        if uploaded_files:
            st.success(f"✅ {len(uploaded_files)} files uploaded")
            
            # Display file list
            with st.expander("📁 Uploaded Files"):
                for i, file in enumerate(uploaded_files, 1):
                    st.write(f"{i}. {file.name} ({file.size} bytes)")
            
            # Process all files button
            if st.button("🔄 Process All PDFs", type="primary"):
                try:
                    # Create progress bars
                    overall_progress = st.progress(0)
                    status_text = st.empty()
                    
                    all_results = []
                    all_translated_texts = []
                    all_dataframes = []
                    total_files = len(uploaded_files)
                    
                    for i, uploaded_file in enumerate(uploaded_files):
                        status_text.text(f"📖 Processing {uploaded_file.name} ({i+1}/{total_files})...")
                        
                        punjabi_text, translated_text, df_table_data = process_single_pdf(
                            uploaded_file, pdf_handler, ocr_processor, translator
                        )
                        
                        if punjabi_text:
                            all_translated_texts.append(f"--- {uploaded_file.name} ---\n{translated_text}")
                        
                        if df_table_data is not None and not df_table_data.empty:
                            all_dataframes.append(df_table_data)
                        else:
                            st.warning(f"⚠️ No structured data extracted for {uploaded_file.name}. It will be excluded from the combined Excel.")
                            # Optionally, add a placeholder to all_dataframes if you need to keep track of failed files in the concat step
                            # For now, we just skip adding it to all_dataframes
                        
                        # Update progress
                        overall_progress.progress((i + 1) / total_files)
                    
                    status_text.text("✅ Batch processing complete!")
                    
                    # Display results
                    st.markdown("---")
                    st.subheader("📊 Batch Processing Results")
                    
                    if all_translated_texts:
                        combined_text = "\n\n".join(all_translated_texts)
                        st.download_button(
                            label="📝 Download Combined Translated Text",
                            data=combined_text,
                            file_name="batch_translations.txt",
                            mime="text/plain"
                        )
                    
                    if all_dataframes:
                        combined_df = pd.concat(all_dataframes, ignore_index=True)
                        st.subheader("📊 Combined Polling Station Data")
                        st.dataframe(combined_df)
                        
                        excel_data = create_excel_file(combined_df)
                        st.download_button(
                            label="📊 Download Combined Excel",
                            data=excel_data,
                            file_name="batch_polling_stations.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.warning("No files were successfully processed.")
                        
                except Exception as e:
                    st.error(f"❌ Error during batch processing: {str(e)}")
                    with st.expander("Show detailed error"):
                        st.code(traceback.format_exc())
    
    # Information section
    with st.expander("ℹ️ How it works"):
        st.markdown("""
        **Single PDF Mode:**
        1. Upload a PDF file containing Punjabi text
        2. OCR extracts text from PDF pages using Tesseract
        3. Text is translated to English using Google Translate
        4. Data is structured into table format with Serial No, Building-Address, Section Covered
        5. Download as text file or Excel spreadsheet
        
        **Batch Processing Mode:**
        1. Upload multiple PDF files at once
        2. Each file is processed individually
        3. Results are combined into a single Excel file
        4. Download combined results or individual file translations
        
        **Supported formats**: PDF files with readable Punjabi text
        **Output formats**: Text files (.txt) and Excel spreadsheets (.xlsx)
        """)
    
    # Usage tips
    with st.expander("💡 Tips for better results"):
        st.markdown("""
        **For better OCR accuracy:**
        - Use PDFs with clear, high-resolution Punjabi text
        - Avoid handwritten text (OCR works best with printed text)
        - Ensure good contrast between text and background
        - Use PDFs with standard fonts
        
        **For better table extraction:**
        - Include keywords like 'building', 'school', 'hall', 'office', 'center', 'ward'
        - Structure text with clear building names and section information
        - Separate different entries on new lines
        """)
    
    # System requirements note
    st.info("📋 The app automatically extracts polling station details and creates structured Excel files with columns: Serial No, Building - Address, Section Covered")

if __name__ == "__main__":
    main()
